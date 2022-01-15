# https://www.marsja.se/your-guide-to-reading-excel-xlsx-files-in-python/
import openpyxl
from pathlib import Path
from IPython.display import display
import pandas as pd
import matplotlib.pyplot as pl
import seaborn as sns
def load_files(folder, file):
    xlsx_file = Path(folder, file)
    wb_obj = openpyxl.load_workbook(xlsx_file)

    # Read the active sheet:
    sheet = wb_obj.active

    # Number os rows and columns
    print("-> ", file, "- Rows: ", sheet.max_row, "Columns: ", sheet.max_column)

    # Build data rows
    tmp_df = pd.DataFrame()
    for i, child in enumerate(sheet.iter_rows(values_only=True)):
        try:
            if i > 0:
                df2 = pd.DataFrame([child], columns=['Year', 'Date', 'Home', 'Score', 'Away', 'Stats', 'Half-time score', 'Match total goals is over 2.5', 'Match total goals', 'Both teams scored'])
                tmp_df = tmp_df.append(df2, ignore_index = True)
        except:
            continue

    return tmp_df

def read_all_files():
    global results
    # Read all the files and concat into a single dataframe
    results = pd.DataFrame()
    xlsx_files = [path for path in Path('data').rglob('*.xlsx')]
    for path in xlsx_files:
        x = str(path).split("/")
        df = load_files(x[0], x[1])
        results = results.append(df, ignore_index=True)

def get_games_history(home):
    tmp_df = pd.DataFrame()
    results_home = results.loc[results['Home'].str.contains(home)]
    results_away = results.loc[results['Away'].str.contains(home)]

    tmp_df = tmp_df.append(results_home, ignore_index=True)
    tmp_df = tmp_df.append(results_away, ignore_index=True)

    return tmp_df

def get_matches_against(home, away):
    tmp_df = pd.DataFrame()

    results_home = results.loc[(results['Home'].str.contains(home)) & (results['Away'].str.contains(away))]
    results_away = results.loc[(results['Away'].str.contains(home)) & (results['Home'].str.contains(away))]

    tmp_df = tmp_df.append(results_home, ignore_index=True)
    tmp_df = tmp_df.append(results_away, ignore_index=True)
    tmp_df = tmp_df.sort_values(by=['Year'])

    return tmp_df


def get_history_matrix(main_team, opponent):
    tmp_df = pd.DataFrame()
    result_list = list()

    # Filter teams
    results_home = results.loc[(results['Home'].str.contains(main_team)) & (results['Away'].str.contains(opponent))]
    results_away = results.loc[(results['Away'].str.contains(main_team)) & (results['Home'].str.contains(opponent))]

    # Merge home and away games
    frames = [results_home, results_away]
    results_merged = pd.concat(frames)
    results_merged = results_merged.sort_values(by=['Year'])

    # Accumulated points
    home_acc_points = 0
    away_acc_points = 0

    # Home games
    for index, row in results_merged.iterrows():
        home_goals = int(row['Score'].split(' - ')[0])
        away_goals = int(row['Score'].split(' - ')[1])
        diff_goals = home_goals - away_goals

        # Calculating points and result (for main team)
        final_result = 'D'
        home_points = 1
        away_points = 1
        if str(row['Home']).strip().lower() == str(main_team).strip().lower():
            if diff_goals < 0:
                home_points = 0
                away_points = 3
                final_result = 'L'
            elif diff_goals > 0:
                home_points = 3
                away_points = 0
                final_result = 'W'
        else:
            if diff_goals > 0:
                home_points = 0
                away_points = 3
                final_result = 'L'
            elif diff_goals < 0:
                home_points = 3
                away_points = 0
                final_result = 'W'

        # Calculating accumulated points
        home_acc_points = home_acc_points + home_points
        away_acc_points = away_acc_points + away_points

        result_list.append(
            [row['Year'], row['Home'], row['Away'], row['Score'], home_points, away_points, home_acc_points,
             away_acc_points, final_result])

    final_df = pd.DataFrame(result_list,
                            columns=['Year', 'Main Team', 'Away', 'Score', 'Main Points', 'Opponent Points',
                                     'Main Acc Points', 'Opponent Acc Points', 'Final Result'])

    return final_df

if __name__ == '__main__':
    read_all_files()
    #print(results.keys())
    #print(results[["Home", "Score", "Away"]])

    pd.set_option('max_columns', None)
    #tmp_df = get_games_history('Vasco da Gama')
    #display(tmp_df)

    #tmp_df = get_matches_against('Vasco da Gama', 'Flamengo')
    #display(tmp_df)

    final_df = get_history_matrix('Vasco da Gama', 'Botafogo')
    display(final_df)





