#!/usr/bin/env python
# coding: utf-8

# In[4]:


# Install a conda package in the current Jupyter kernel
import sys
get_ipython().system('conda install --yes --prefix {sys.prefix} openpyxl')


# In[13]:


import openpyxl
from pathlib import Path
from IPython.display import display
import pandas as pd


# In[73]:


def load_files(folder, file):
    xlsx_file = Path(folder, file)
    wb_obj = openpyxl.load_workbook(xlsx_file)

    # Read the active sheet:
    sheet = wb_obj.active

    # Number os rows and columns
    print("-> ", file, "- Rows: ", sheet.max_row, "Columns: ", sheet.max_column)

    # Build data rows
    tmp_df = pd.DataFrame()
    for i, child in enumerate(sheet.iter_rows(values_only=True)):
        try:
            if i > 0:
                df2 = pd.DataFrame([child], columns=['Year', 'Date', 'Home', 'Score', 'Away', 'Stats', 'Half-time score', 'Match total goals is over 2.5', 'Match total goals', 'Both teams scored'])
                tmp_df = tmp_df.append(df2, ignore_index = True)
        except:
            continue

    return tmp_df

def read_all_files():
    global results
    # Read all the files and concat into a single dataframe
    results = pd.DataFrame()
    xlsx_files = [path for path in Path('data').rglob('*.xlsx')]
    for path in xlsx_files:
        x = str(path).split("/")
        df = load_files(x[0], x[1])
        results = results.append(df, ignore_index=True)

def get_games_history(home):
    tmp_df = pd.DataFrame()
    results_home = results.loc[results['Home'].str.contains(home)]
    results_away = results.loc[results['Away'].str.contains(home)]

    tmp_df = tmp_df.append(results_home, ignore_index=True)
    tmp_df = tmp_df.append(results_away, ignore_index=True)

    return tmp_df

def get_matches_against(home, away):
    tmp_df = pd.DataFrame()

    results_home = results.loc[(results['Home'].str.contains(home)) & (results['Away'].str.contains(away))]
    results_away = results.loc[(results['Away'].str.contains(home)) & (results['Home'].str.contains(away))]

    tmp_df = tmp_df.append(results_home, ignore_index=True)
    tmp_df = tmp_df.append(results_away, ignore_index=True)
    tmp_df = tmp_df.sort_values(by=['Year'])

    return tmp_df


# In[112]:


read_all_files()


# In[79]:


pd.set_option('max_columns', None)
tmp_df = get_matches_against('Vasco da Gama', 'Flamengo')
tmp_df = tmp_df.sort_values(by=['Year'])
display(tmp_df)


# In[146]:


def get_history_matrix(main_team, opponent):
    tmp_df = pd.DataFrame()
    result_list = list() 

    # Filter teams
    results_home = results.loc[(results['Home'].str.contains(main_team)) & (results['Away'].str.contains(opponent))]
    results_away = results.loc[(results['Away'].str.contains(main_team)) & (results['Home'].str.contains(opponent))]
    
    # Merge home and away games
    frames = [results_home, results_away]
    results_merged = pd.concat(frames)
    results_merged = results_merged.sort_values(by=['Year'])
    
    # Accumulated points
    home_acc_points = 0
    away_acc_points = 0
    
    # Home games
    for index, row in results_merged.iterrows():
        home_goals = int(row['Score'].split(' - ')[0])
        away_goals = int(row['Score'].split(' - ')[1])
        diff_goals = home_goals - away_goals
        
        # Calculating points and result (for main team)
        final_result = 'D'
        home_points = 1
        away_points = 1
        if str(row['Home']).strip().lower() == str(main_team).strip().lower():
            if diff_goals<0:
                home_points = 0
                away_points = 3
                final_result = 'L'
            elif diff_goals>0:
                home_points = 3
                away_points = 0
                final_result = 'W'
        else:
            if diff_goals>0:
                home_points = 0
                away_points = 3
                final_result = 'L'
            elif diff_goals<0:
                home_points = 3
                away_points = 0
                final_result = 'W'

        # Calculating accumulated points
        home_acc_points = home_acc_points + home_points
        away_acc_points = away_acc_points + away_points
        
        result_list.append([row['Year'], row['Home'], row['Away'], row['Score'], home_points, away_points, home_acc_points, away_acc_points, final_result])


    final_df = pd.DataFrame(result_list, columns = ['Year', 'Main Team', 'Away', 'Score', 'Main Points', 'Opponent Points', 'Main Acc Points', 'Opponent Acc Points', 'Final Result'])
        
    return final_df


# In[147]:


final_df = get_history_matrix('Vasco da Gama', 'Botafogo')
display(final_df)


# In[148]:


final_df[['Year', 'Main Acc Points', 'Opponent Acc Points']].plot.line(x='Year')


# In[210]:


def get_history_matrix(main_team):
    tmp_df = pd.DataFrame()
    result_list = list() 

    # Filter teams
    results_home = results.loc[(results['Home'].str.contains(main_team))]
    results_away = results.loc[(results['Away'].str.contains(main_team))]
    
    # Merge home and away games
    frames = [results_home, results_away]
    results_merged = pd.concat(frames)
    results_merged = results_merged.sort_values(by=['Year'])
    
    # Accumulated points
    home_acc_points = 0
    away_acc_points = 0
    
    # Home games
    for index, row in results_merged.iterrows():
        home_goals = int(row['Score'].split(' - ')[0])
        away_goals = int(row['Score'].split(' - ')[1])
        diff_goals = home_goals - away_goals
        
        # Calculating points and result (for main team)
        final_result = 'D'
        home_points = 1
        away_points = 1
        if str(row['Home']).strip().lower() == str(main_team).strip().lower():
            if diff_goals<0:
                home_points = 0
                away_points = 3
                final_result = 'L'
            elif diff_goals>0:
                home_points = 3
                away_points = 0
                final_result = 'W'
        else:
            if diff_goals>0:
                home_points = 0
                away_points = 3
                final_result = 'L'
            elif diff_goals<0:
                home_points = 3
                away_points = 0
                final_result = 'W'

        # Calculating accumulated points
        home_acc_points = home_acc_points + home_points
        away_acc_points = away_acc_points + away_points
        
        result_list.append([row['Year'], row['Home'], row['Away'], row['Score'], home_points, away_points, home_acc_points, away_acc_points, final_result])


    final_df = pd.DataFrame(result_list, columns = ['Year', 'Main Team', 'Away', 'Score', 'Main Points', 'Opponent Points', 'Main Acc Points', 'Opponent Acc Points', 'Final Result'])
        
    return final_df


# In[212]:


final_global_df = get_history_matrix('Vasco da Gama')
display(final_global_df)


# In[167]:


from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score , confusion_matrix


# In[257]:


x_data = final_global_df.iloc[:,4:5]
y_data = final_global_df['Main Points']


# In[258]:


x_train, x_test, y_train, y_test = train_test_split(x_data, y_data, test_size = 0.33, random_state = 12)
model = LogisticRegression(solver='lbfgs', max_iter=10000)
model.fit(x_train, y_train)

prediction = dict()
predictions = model.predict(x_test)

# Use score method to get accuracy of model
score = model.score(x_test, y_test)
print("Accuracy:", score)
print(model.coef_)

conf_mat_logist = confusion_matrix(y_test, predictions)
print('Logist \r', conf_mat_logist)


# In[259]:


import matplotlib.pyplot as plt
import seaborn as sns

plt.figure(figsize=(9,9))
sns.heatmap(conf_mat_logist, annot=True, fmt=".3f", linewidths=.5, square = True, cmap = 'Blues_r');
plt.ylabel('Pontos Perdidos');
plt.xlabel('Pontos Ganhos');
all_sample_title = 'Accuracy Score: {0}'.format(score)
plt.title(all_sample_title, size = 15);


# In[264]:


print("Predicted (Global) = %s" % (predictions[0]))


# In[229]:


final_global_df['Main Points'].describe()


# In[ ]:




